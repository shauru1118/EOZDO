import openpyxl as pxl
import os
import EosdoRequest

b_ = ["Сотрудник", "Период", "Выполнено задач", "Просрочено задач", "", "", "", "", "", "", "",]
columns = {
    0:"A",
    1:"B",
    2:"C",
    3:"D",
    4:"E",
    5:"F",
    6:"G",
    7:"H",
    8:"I",
    9:"J",
    10:"K",
    11:"L",
}


class ExelWB:

    def __init__(self):

        self.filename = "premia.xlsx"
        self.author = "Игорь"
        self.employes = [
            "Андросов",
            "Белов",
            "Васильев",
            "Горбачев",
            "Демидов",
            "Ефремов",
            "Жуков",
            "Зайцев",
            "Иванов",
            "Карпов",
            "Кириллов",
            "Левов",
            "Михайлов",
            "Николаев",
            "Орлов",
            "Петров",
        ]
        self.data = {
            self.employes[0]: [],
            self.employes[1]: [],
            self.employes[2]: [],
            self.employes[3]: [],
            self.employes[4]: [],
            self.employes[5]: [],
            self.employes[6]: [],
            self.employes[7]: [],
            self.employes[8]: [],
            self.employes[9]: [],
            self.employes[10]: [],
            self.employes[11]: [],
            self.employes[12]: [],
            self.employes[13]: [],
            self.employes[14]: [],
            self.employes[15]: [],
        }

        self.wb = pxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Premia"
        print(f"Создан эксель файл: {self.filename}")

        self.ws["A1"] = f"Автор: {self.author}"

        for i in range(len(b_)):
            self.ws[f"{columns[i]}3"] = b_[i]
        for i in range(len(self.employes)):
            self.ws[f"A{i+4}"] = self.employes[i]

        self.eosdo = EosdoRequest.EOSDO()
        
    def get_data(self):
        # Заполняем данными
        

        print(f"Данные получины")

    def save_data(self):
        # Сохраняем данные в эксель файле

        print(f"Данные сохранены в эксель файл: {self.filename}")

    def show(self):
        self.wb.save(self.filename)
        os.startfile(self.filename)

