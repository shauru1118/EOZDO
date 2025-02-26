import openpyxl as xl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from rich.progress import Progress

from copy import copy
from time import sleep

import xlrd

format_empl_file = "C:\\AlbertsProgram\\ess.xlsx"
rich_delay = 0.010

yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
green_fill = PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid")
emploe_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
 
emploe1_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid") # yellow
emploe2_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid") # green

itog_fill = PatternFill(start_color="99FFFF", end_color="99FFFF", fill_type="solid")

red_font = Font(color="FF0000")
yellow_font = Font(color="666600")
green_font = Font(color="006600")


black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # black
black_font = Font(color="000000", bold=True)
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # white
white_font = Font(color="FFFFFF", bold=True)

border_style = Border(
    left=Side(border_style="thin", color="000000"),  # Линия слева
    right=Side(border_style="thin", color="000000"),  # Линия справа
    top=Side(border_style="thin", color="000000"),  # Линия сверху
    bottom=Side(border_style="thin", color="000000")  # Линия снизу
)

a_width = 35
bcd_width = 15
e_width = 15


class ExelWB:
    def __init__(self, name=None, res_file_name=None, temp_table_name: dict = None):
        if name != None:
            self.wb = xlrd.open_workbook(name)
            
        # print(f"xlrd: Workbook opened: '{name}'")

        self.author = "Игорь"
        self.filename = name
        self.res_file_name = res_file_name
        self.res_premia_wb = xl.load_workbook(self.res_file_name)
        self.svod_table = temp_table_name
        self.emploes = {}
        self.emploe_switch = 1
        
        self.format_emploeers = []
        emp = xl.load_workbook(format_empl_file)
        format_sheet = emp.active
        for i in range(format_sheet.max_row):
            self.format_emploeers.append(format_sheet.cell(row=i+1, column=1).value)
            
        self.progress = Progress()
        self.progress.start()
        
    
    def do_prosrochki(self):

        sheet = self.wb.sheet_by_index(0)

        all_pros = {}
        month_pros = {}

        employees = [sheet.cell_value(row, 11) for row in range(2, sheet.nrows-1)] # col11
        dates = [int(sheet.cell_value(row, 8).split(".")[1]) for row in range(2, sheet.nrows-1)] # col8
        status = [sheet.cell_value(row, 4).split(", ") for row in range(2, sheet.nrows-1)]

        pros_task = self.progress.add_task("[cyan] ПросрАчки...      ", total=len(employees))
        for i in range(len(employees)):
            cell = employees[i]
            # print("task: ", end="")
            names = cell.split("\n")
            for name in names:
                if len(name) == 0:
                    continue
                try:
                    arr = name.split("/")
                    # print("\tname:", arr[0].strip(), "\n\t\tstatus:", arr[1].strip())
                    employee = arr[0].strip()
                    if ("но не принят" in status[i]) and ("Отчет отклонен на доработку" in arr):
                        if employee not in self.svod_table.keys():
                            self.svod_table[employee] = {"y": 1}
                        else:
                            if "y" not in self.svod_table[employee].keys():
                                self.svod_table[employee]["y"] = 1

                            else:
                                self.svod_table[employee]["y"] += 1

                except:
                    # print("\tname:", name, "\n\t\tstatus: NONE")
                    employee = name.strip()

                if employee not in all_pros.keys():
                    all_pros[employee] = 1
                else:
                    all_pros[employee] += 1

                if employee not in self.svod_table.keys():
                    self.svod_table[employee] = {"r": 1}
                else:
                    if "r" not in self.svod_table[employee].keys():
                        self.svod_table[employee]["r"] = 1
                    else:
                        self.svod_table[employee]["r"] += 1
                
                if employee not in month_pros.keys():
                    month_pros[employee] = {dates[i]: 1}
                else:
                    if dates[i] not in month_pros[employee].keys():
                        month_pros[employee][dates[i]] = 1
                    else:
                        month_pros[employee][dates[i]] += 1
            # print()
            
            self.progress.update(pros_task, advance=1)
            # print('.')
            sleep(rich_delay)
        # sleep(10)
        # Sorting by problems

        sorted_all_pros = dict(sorted(all_pros.items(), key=lambda item: item[1], reverse=True))
        sorted_month_pros = {}
        for employee, counts in sorted_all_pros.items():
            sorted_month_pros[employee] = month_pros[employee]
        # sorting by months
        for employee, dates in sorted_month_pros.items():
            sorted_month_pros[employee] = dict(sorted(dates.items(), key=lambda item: item[0], reverse=True))

        # print("\n\n\nAll Problems:")

        # for employee, count in sorted_all_pros.items():
        #     print(employee, " - ", count)


        # print("\n\n\nMonth Problems:")

        # for employee, dates in sorted_month_pros.items():
        #     print(employee)
        #     for date, count in dates.items():
        #         print("\t", date, " - ", count)
        
        res_sheet = self.res_premia_wb.active
        res_sheet.title = "Просрочки и закрыто"
        res_sheet.cell(row=1, column=1).value = "Просрочки"
        res_sheet.merge_cells("A1:C1")
        res_sheet.cell(row=1, column=1).fill = black_fill
        res_sheet.cell(row=1, column=1).font = white_font
        res_sheet.cell(row=3, column=1).value = "Сотрудник"
        res_sheet.cell(row=3, column=1).fill = emploe_fill
        res_sheet.cell(row=res_sheet.max_row, column=1).font = black_font
        res_sheet.cell(row=3, column=2).value = "Месяц"
        res_sheet.cell(row=3, column=2).fill = yellow_fill
        res_sheet.cell(row=res_sheet.max_row, column=2).font = black_font
        res_sheet.cell(row=3, column=3).value = "Количество\nпрострочек\nза месяц"
        res_sheet.cell(row=3, column=3).fill = red_fill
        res_sheet.cell(row=res_sheet.max_row, column=3).font = black_font

        for employee, dates in sorted_month_pros.items():

            if self.emploe_switch == 1:

                res_sheet.cell(row=res_sheet.max_row + 2, column=1).value = employee

                res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe1_fill

                res_sheet.cell(row=res_sheet.max_row, column=2).value = "Итого"
                res_sheet.cell(row=res_sheet.max_row, column=2).fill = emploe1_fill
                
                res_sheet.cell(row=res_sheet.max_row, column=3).value = sorted_all_pros[employee]
                res_sheet.cell(row=res_sheet.max_row, column=3).fill = emploe1_fill

                for date, count in dates.items():
                    res_sheet.cell(row=res_sheet.max_row+1, column=2).value = date
                    res_sheet.cell(row=res_sheet.max_row, column=3).value = count
                    res_sheet.cell(row=res_sheet.max_row, column=3).fill = emploe1_fill
                    res_sheet.cell(row=res_sheet.max_row, column=2).fill = emploe1_fill

                    # res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe1_fill

                res_sheet.row_dimensions[res_sheet.max_row+1].height = 1    
                
                self.emploe_switch = 2
                
            elif self.emploe_switch == 2:
                res_sheet.cell(row=res_sheet.max_row + 2, column=1).value = employee

                res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe2_fill

                res_sheet.cell(row=res_sheet.max_row, column=2).value = "Итого"
                res_sheet.cell(row=res_sheet.max_row, column=2).fill = emploe2_fill
                
                res_sheet.cell(row=res_sheet.max_row, column=3).value = sorted_all_pros[employee]
                res_sheet.cell(row=res_sheet.max_row, column=3).fill = emploe2_fill

                for date, count in dates.items():
                    res_sheet.cell(row=res_sheet.max_row+1, column=2).value = date
                    res_sheet.cell(row=res_sheet.max_row, column=3).value = count
                    res_sheet.cell(row=res_sheet.max_row, column=3).fill = emploe2_fill
                    res_sheet.cell(row=res_sheet.max_row, column=2).fill = emploe2_fill

                    # res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe2_fill

                res_sheet.row_dimensions[res_sheet.max_row+1].height = 1

                self.emploe_switch = 1

        # for col in res_sheet.columns:
        #     max_length = 0
        #     col_letter = col[-1].column_letter
        #     for cell in col:
        #         if cell.value:
        #             max_length = max(max_length, len(str(cell.value)))
        #         cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        #     res_sheet.column_dimensions[col_letter].width = max_length + 2

        # for row in range(1, res_sheet.max_row + 1):
        #     res_sheet.row_dimensions[row].height = 22
        # res_sheet.row_dimensions[3].height = 50
                
        self.res_premia_wb.save(self.res_file_name)

        return self.svod_table

    
    def do_vipolneno(self):
        sheet = self.wb.sheet_by_index(0)
        
        all_vipo = {}

        employees = [sheet.cell_value(row, 10) for row in range(2, sheet.nrows-1)] # col10

        vip_task = self.progress.add_task("[cyan] Выполнено...      ", total=len(employees))
        for i in range(len(employees)):
            cell = employees[i]
            # print("task: ", end="")
            names = cell.split("\n")
            for name in names:
                if len(name) == 0:
                    continue
                try:
                    arr = name.split("/")
                    # print("\tname:", arr[0].strip(), "\n\t\tstatus:", arr[1].strip())
                    employee = arr[0].strip()
                except:
                    # print("\tname:", name, "\n\t\tstatus: NONE")
                    employee = name.strip()

                if employee not in all_vipo.keys():
                    all_vipo[employee] = 1
                else:
                    all_vipo[employee] += 1
                
                if employee not in self.svod_table.keys():
                    self.svod_table[employee] = {"g": 1}
                else:
                    if "g" not in self.svod_table[employee].keys():
                        self.svod_table[employee]["g"] = 1
                    else:
                        self.svod_table[employee]["g"] += 1
            # print()
            self.progress.update(vip_task, advance=1)
            sleep(rich_delay/2.5)
        
        
        sorted_all_vipo = dict(sorted(all_vipo.items(), key=lambda item: item[1], reverse=True))

        # print("\n\n\nSorted by problems:")

        # for employee, count in sorted_all_vipo.items():
            # print(employee, " - ", count)

        # append to excel


        res_sheet = self.res_premia_wb.active

        res_sheet.cell(row=res_sheet.max_row+3, column=1).value = "Выполенные"
        res_sheet.merge_cells(f"A{res_sheet.max_row}:B{res_sheet.max_row}")
        res_sheet.cell(row=res_sheet.max_row, column=1).font = white_font
        res_sheet.cell(row=res_sheet.max_row, column=1).fill = black_fill

        res_sheet.cell(row=res_sheet.max_row+2, column=1).value = "Сотрудник"
        res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe_fill
        res_sheet.cell(row=res_sheet.max_row, column=1).font = black_font
        
        res_sheet.cell(row=res_sheet.max_row, column=2).value = "Всего\nвыполенных"
        res_sheet.cell(row=res_sheet.max_row, column=2).fill = green_fill
        res_sheet.cell(row=res_sheet.max_row, column=2).font = black_font
        
        title_row = res_sheet.max_row
        res_sheet.row_dimensions[res_sheet.max_row].height = 40

        for employee, count in sorted_all_vipo.items():
            res_sheet.cell(row=res_sheet.max_row + 1, column=1).value = employee
            res_sheet.cell(row=res_sheet.max_row, column=1).fill = emploe_fill
        
            res_sheet.cell(row=res_sheet.max_row, column=2).value = count
            res_sheet.cell(row=res_sheet.max_row, column=2).fill = green_fill


        for col in res_sheet.columns:
            max_length = 0
            col_letter = col[-1].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    cell.border = border_style
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            res_sheet.column_dimensions[col_letter].width = max_length + 2


        for row in range(1, res_sheet.max_row + 1):
            res_sheet.row_dimensions[row].height = 25

        res_sheet.row_dimensions[3].height = 50
        res_sheet.row_dimensions[title_row].height = 40
        res_sheet.column_dimensions["B"].width = 15
        res_sheet.column_dimensions["c"].width = 20

        self.res_premia_wb.save(self.res_file_name)

        return self.svod_table

    def do_svod(self):

        # print("\n\n\nСводная таблица:", self.svod_table)
        

        res_sheet = self.res_premia_wb.create_sheet("Таблица")

        sheet_to_move = self.res_premia_wb["Таблица"]
        self.res_premia_wb._sheets.remove(sheet_to_move)  # Удаляем из списка
        self.res_premia_wb._sheets.insert(0, sheet_to_move)  # Вставляем на нужное место

        res_sheet = self.res_premia_wb["Таблица"]

        svod_task = self.progress.add_task("[cyan] Сводная таблица...", total=len(self.svod_table.keys())*2)
        
        res_sheet.cell(row=1, column=1).value = "Сводная таблица"
        res_sheet.cell(row=1, column=1).font = Font(color="FFFFFF")
        res_sheet.cell(row=1, column=1).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        res_sheet.merge_cells("A1:E1")
    
        res_sheet.cell(row=3, column=1).value = "Поручения"
        res_sheet.cell(row=3, column=1).fill = itog_fill
        res_sheet.merge_cells("A3:E3")
        # res_sheet["B3"].alignment = Alignment(horizontal="center", vertical="center")
        res_sheet.cell(row=4, column=1).value = "Сотрудник"
        res_sheet.cell(row=4, column=1).fill = emploe_fill
        res_sheet.cell(row=4, column=1).font = Font(bold=True)
        res_sheet.cell(row=4, column=2).value = "Просрочено,\nотчет отклонен"
        res_sheet.cell(row=4, column=2).fill = red_fill
        res_sheet.cell(row=4, column=2).font = Font(bold=True)
        res_sheet.cell(row=4, column=3).value = "Отчет на проверке"
        res_sheet.cell(row=4, column=3).fill = yellow_fill
        res_sheet.cell(row=4, column=3).font = Font(bold=True)
        res_sheet.cell(row=4, column=4).value = "Выполнено"
        res_sheet.cell(row=4, column=4).fill = green_fill
        res_sheet.cell(row=4, column=4).font = Font(bold=True)
        res_sheet.cell(row=4, column=5).value = "Итог по\nсотруднику"
        res_sheet.cell(row=4, column=5).fill = itog_fill
        res_sheet.cell(row=4, column=5).font = Font(bold=True)
        
        self.svod_table = dict(sorted(self.svod_table.items()))
        
        svod_table_format = {}
        svod_table_other = {}
        
        for employee in self.format_emploeers:
            right_employee = next((e for e in self.svod_table if employee.split()[0] in e), None)
            # print(self.format_emploeers.index(employee), right_employee)
            try: svod_table_format[right_employee] = self.svod_table[right_employee]
            except: pass
            self.svod_table.pop(right_employee, None)
            
        svod_table_other = copy(self.svod_table)
        
        self.svod_table.clear()
        
        for e, d in svod_table_format.items():
            self.svod_table[e] = d
        
        for e, d in svod_table_other.items():
            self.svod_table[e] = d
        
            
        for employee, done in self.svod_table.items():
            e = res_sheet.cell(row=res_sheet.max_row + 1, column=1)
            e.value = employee
            e.fill = emploe_fill
            r = res_sheet.cell(row=res_sheet.max_row, column=2)
            r.value = done.get("r", 0)
            r.font = red_font
            r.fill = red_fill
            
            y = res_sheet.cell(row=res_sheet.max_row, column=3)
            if done.get("y", 0) != 0:
                y.value = done.get("y", 0)
            y.font = yellow_font
            y.fill = yellow_fill
            
            g = res_sheet.cell(row=res_sheet.max_row, column=4)
            g.value = done.get("g", 0)
            g.font = green_font
            g.fill = green_fill
            
            itog_res = res_sheet.cell(row=res_sheet.max_row, column=5)
            itog_res.value = done.get("r", 0) + done.get("y", 0) + done.get("g", 0)
            itog_res.fill = itog_fill
            
            self.progress.update(svod_task, advance=1)
            sleep(rich_delay)

        temp = res_sheet.cell(row=res_sheet.max_row+1, column=1)
        temp.value = "Итого"
        temp.fill = itog_fill
        r = []
        y = []
        g = []
        for val in self.svod_table.values():
            try:
                r.append(val.get("r", 0))
                y.append(val.get("y", 0))
                g.append(val.get("g", 0))
            except:
                pass
            self.progress.update(svod_task, advance=1)
            sleep(rich_delay)


        r_res = res_sheet.cell(row=res_sheet.max_row, column=2)
        r_res.value = sum(r)
        temp = copy(red_font); temp.bold = True; r_res.font = temp
        r_res.fill = itog_fill
        
        y_res = res_sheet.cell(row=res_sheet.max_row, column=3)
        y_res.value = sum(y)
        temp = copy(yellow_font); temp.bold = True; y_res.font = temp
        y_res.fill = itog_fill
        
        g_res = res_sheet.cell(row=res_sheet.max_row, column=4)
        g_res.value = sum(g)
        temp = copy(green_font); temp.bold = True; g_res.font = temp
        g_res.fill = itog_fill
        
        res_res = res_sheet.cell(row=res_sheet.max_row, column=5)
        res_res.value = sum(r) + sum(y) + sum(g)
        res_res.font = Font(bold=True, color="FFFFFF")
        res_res.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        for col in res_sheet.columns:
            max_length = 0
            col_letter = col[-1].column_letter
            for cell in col:
                if (cell.value or cell.value == 0) and cell.row != 2:
                    max_length = max(max_length, len(str(cell.value)))
                    # lines = str(cell.value).count("\n") + 1  # Считаем строки текста
                    # res_sheet.row_dimensions[cell.row].height = lines * 50  # 15 пикселей на строку
                    # cell.border = border_style
                cell.border = border_style
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # Учитываем перенос строк
            # res_sheet.column_dimensions[col_letter].width = max_length + 10

        res_sheet.column_dimensions["A"].width = a_width
        res_sheet.column_dimensions["B"].width = bcd_width
        res_sheet.column_dimensions["C"].width = bcd_width
        res_sheet.column_dimensions["D"].width = bcd_width
        res_sheet.column_dimensions["E"].width = e_width


        for row in range(1, res_sheet.max_row + 1):
            res_sheet.row_dimensions[row].height = 22
        res_sheet.row_dimensions[4].height = 30
        

        self.res_premia_wb.save(self.res_file_name)
        
    def create_res_premia(res_file_name: str):
        xl.Workbook().save(res_file_name)

    

    def end(self):
        self.res_premia_wb.close()
        self.progress.stop()